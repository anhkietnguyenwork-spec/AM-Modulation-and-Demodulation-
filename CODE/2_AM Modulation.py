import numpy as np
import sounddevice as sd
import os
from tkinter import Tk, filedialog
from wav_reader import read_wav
import soundfile as sf
from openpyxl import load_workbook
from dsp_utils import plot_fft, plot_spectrogram


# ==================== CHỌN FILE ==================== 
root = Tk(); root.withdraw()
file_path = filedialog.askopenfilename(
    title="Chọn file _FFT.wav",
    filetypes=[("WAV","*_FFT.wav")]
)
if not file_path: exit("LỖI: Chưa chọn file _FFT.wav")


# ==================== TÊN FILE EXCEL TƯƠNG ỨNG ====================
xlsx_path = file_path.replace("_FFT.wav", ".xlsx")   # Cùng tên – đổi đuôi
if not os.path.exists(xlsx_path):
    exit(f"LỖI: Không tìm thấy file Excel: {xlsx_path}")


# ==================== ĐỌC THÔNG TIN TỪ EXCEL ====================
wb = load_workbook(xlsx_path)
ws = wb.active

fs = ws["A2"].value 
if fs is None:
    exit("LỖI: Ô A2 chưa có Fs => Phải chạy FFT trước!")

print(f"\n Đọc Excel - Sampling Fs = {fs} Hz")


# ==================== USER NHẬP THAM SỐ → GHI VÀO EXCEL ====================
print("\n Nhập tham số AM (dùng tiếp cho các bước sau):")
f_c = float(input("Tần số sóng mang fc (Hz) = "))
k_a = float(input("Hệ số điều chế k (0 < k ≤ 1) = "))
if not (0 < k_a <= 1):
    exit("LỖI: k phải nằm trong (0,1]")

ws["C2"] = f_c
ws["D2"] = k_a
wb.save(xlsx_path)

print(f"\n>> Đã cập nhật Excel:")
print(f"A2 = {fs}     (Tần số lấy mẫu gốc)")
print(f"C2 = {f_c} Hz (Tần số sóng mang AM)")
print(f"D2 = {k_a}    (Hệ số điều chế)")


# ==================== LOAD WAV = Fs từ Excel ====================
y, _ = read_wav(file_path, fs)     # tần số lấy mẫu Fs từ file Excel
if y.ndim > 1: y=y[:,0]
y = y.astype(float)
y /= np.max(np.abs(y))
N=len(y)
t=np.arange(N)/fs


# ==================== AM MODULATION ====================
carrier=np.cos(2*np.pi*f_c*t)
y_am = (1 + k_a*y) * carrier
y_am/=np.max(np.abs(y_am))

# ==================== Playback lựa chọn ====================
print("\n--------------- PHÁT ÂM THANH ----------------")
print("1 = Nghe tín hiệu sau AM")
print("0 = Bỏ qua phát thử")
choice = input("Chọn cách phát âm thanh: ")
if choice=="1": 
    print(">> Phát tín hiệu gốc...")
    sd.play(y_am,fs); sd.wait()


# ==================== PLAY + SAVE ====================
out = file_path.replace("_FFT.wav","_AM.wav")
sf.write(out,y_am.astype("float32"),fs)
print("\n>> Xuất file:", out)


# ========= HIỂN THỊ PHỔ FFT =========
f_plot_max = float(input("\nNhập tần số cần hiển thị FFT đến (Hz): "))
Y=np.fft.fft(y)
Y_am=np.fft.fft(y_am)
freqs=np.fft.fftfreq(N,1/fs)
idx=np.where((freqs>=0)&(freqs<=f_plot_max*2))

plot_fft(
    freqs[idx], Y[idx], Y_am[idx],
    fs=fs, cutoff=f_c,
    label1="Trước AM",
    label2=f"Sau AM (fc={f_c}Hz)",
    title=f"So sánh trước và sau AM: 0 → {f_plot_max} Hz"
)

plot_spectrogram(
    y, y_am, fs,
    fmax=f_plot_max,
    title1="Trước AM",
    title2="Sau AM",
    suptitle=f"Phổ sau AM (fc={f_c})"
)
