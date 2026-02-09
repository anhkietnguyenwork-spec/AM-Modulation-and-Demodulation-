import numpy as np
import sounddevice as sd
import os
from tkinter import Tk, filedialog
from wav_reader import read_wav
from scipy.io.wavfile import write
from dsp_utils import plot_fft, plot_spectrogram
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.signal import resample_poly 


# ================= CHỌN FILE =================
root = Tk(); root.withdraw()
file_path = filedialog.askopenfilename(
    title="Chọn file tín hiệu AM (_AM.wav)",
    filetypes=[("WAV","*_AM.wav")]
)
if not file_path: exit("LỖI: Chưa chọn file!")


# ============ LẤY Fs GỐC TỪ EXCEL =============
xlsx = file_path.replace("_AM.wav",".xlsx")
if not os.path.exists(xlsx): exit("LỖI: Không thấy file Excel!")

wb = load_workbook(xlsx)
ws = wb.active
initial_fs = ws["A2"].value
if initial_fs is None: exit("LỖI: Chưa có Fs (ô A2 trống)")

print(f"\nFs gốc từ Excel = {initial_fs} Hz")


# =============== LOAD WAV GIỮ NGUYÊN FS ===============
y,_ = read_wav(file_path,int(initial_fs))
if y.ndim>1: y=y[:,0]
y=y.astype(float); y/=np.max(np.abs(y))
N=len(y)
print(f"Đã load file AM — Fs={initial_fs}Hz | {N} mẫu")


# ================== USER NHẬP FS MỚI ==================
fs_new=float(input("\nNhập tần số sampling mới fs' (Hz): "))
if fs_new<=0: exit("LỖIfs_new phải >0")


# ============ RESAMPLE ============
ratio = fs_new/initial_fs
y_new = resample_poly(y, fs_new, initial_fs)

print(f"Resampled thành công — từ {initial_fs}Hz => {fs_new} Hz: {len(y_new)} mẫu")


# ===== Ghi vào ô E2 file Excel =====
ws["E2"] = fs_new
wb.save(xlsx)
print(f"Lưu E2 = {fs_new} Hz")


# =============== PLAY + SAVE =================
sd.play(y_new,int(fs_new)); sd.wait()
out = file_path.replace("_AM.wav","_SAMP.wav")
write(out,int(fs_new),(y_new*32767).astype(np.int16))
print(f"\n💾 Saved → {out}")

# tạo timeline tương ứng
t_old = np.arange(len(y)) / initial_fs
t_new = np.arange(len(y_new)) / fs_new

ms = 0.005  
n_old = int(ms * initial_fs)
n_new = int(ms * fs_new)

plt.figure(figsize=(12,4))
plt.plot(t_old[:n_old], y[:n_old], label=f"Tín hiệu gốc {initial_fs}Hz", linewidth=1.2)
plt.plot(t_new[:n_new], y_new[:n_new], label=f"Tín hiệu sau lấy mẫu {fs_new}Hz", linewidth=1.1)

plt.title("Đồ thị theo miền thời gian")
plt.xlabel("Thời gian (s)")
plt.ylabel("Biên độ")
plt.grid(True)
plt.legend()
plt.tight_layout()
plt.show()
# ============================================================
