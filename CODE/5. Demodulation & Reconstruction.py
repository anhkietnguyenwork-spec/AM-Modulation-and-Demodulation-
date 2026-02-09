import numpy as np
import sounddevice as sd
import os
from tkinter import Tk, filedialog
from wav_reader import read_wav
from scipy.io.wavfile import write
from scipy.signal import firwin, lfilter
from dsp_utils import plot_fft, plot_spectrogram
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# ====================== SELECT FILE ======================
root = Tk(); root.withdraw()
file_path = filedialog.askopenfilename(
    title="Chọn tín hiệu lượng tử (_QNT.wav)",
    filetypes=[("WAV","*_QNT.wav")]
)
if not file_path:
    exit("LỖI: Chưa chọn file _QNT.wav !")


# ====================== LOAD EXCEL DATA ======================
xlsx = file_path.replace("_QNT.wav", ".xlsx")
if not os.path.exists(xlsx):
    exit("LỖI: Không có file Excel đính kèm — cần từ bước trước!")

wb = load_workbook(xlsx)
ws = wb.active

# Lấy dữ liệu đúng ô theo yêu cầu
fs     = ws["E2"].value   # sampling sau bước Sampling
fc     = ws["C2"].value   # sóng mang AM
f_base = ws["B2"].value   # cutoff FFT (băng thông baseband gốc)
bits_q = ws["F2"].value   # số bit lượng tử hóa

if fs is None or fc is None or f_base is None:
    exit("⚠ Lỗi thiếu dữ liệu Excel (B2/C2/E2/F2)! Kiểm tra file trước đó.")

# ====== CHỈNH CÁCH CHỌN BĂNG THÔNG LPF SAU DEMOD ======
# Muốn giữ gần như toàn bộ baseband sau FFT → cho băng thông ≈ f_base,
# nhưng vẫn đảm bảo < fc và < Nyquist.
bw = f_base

# Nếu người dùng chọn f_base ≥ fc thì thu hẹp nhẹ để tránh chồng phổ
if bw > 0.8 * fc:
    print("f_base gần / lớn hơn fc → tự thu hẹp baseband còn 0.8*fc để tránh chồng phổ.")
    bw = 0.8 * fc

# Không cho vượt quá 0.45*fs (an toàn Nyquist)
bw = min(bw, 0.45 * fs)

print("\n========== LOADED PARAMETERS FROM EXCEL ==========")
print(f"Tần số lấy mẫu Fs      = {fs} Hz  (E2)")
print(f"Tần số sóng mang Fc    = {fc} Hz  (C2)")
print(f"Tần số cắt FFT         = {f_base} Hz")
print(f"LPF baseband dùng      = {bw:.1f} Hz  (≈ băng gốc sau FFT)")
print(f"Số bit lượng tử        = {bits_q} bit  (F2)")
print("=================================================")


# ====================== READ WAV ======================
y_q, _ = read_wav(file_path, int(fs))  # dùng đúng Fs từ file excel!
if y_q.ndim > 1:
    y_q = y_q[:, 0]

y_q = y_q.astype(float)
y_q /= np.max(np.abs(y_q))
N = len(y_q)
t = np.arange(N) / fs

print(f"\nĐã tải xong file _QNT → Fs={fs} Hz | {N} mẫu")


# ====================== DIGITAL DEMOD ======================
carrier = np.cos(2*np.pi*fc*t)
y_demod = y_q * carrier

# ===== FIR LPF: chỉ giữ baseband, loại fc & 2fc =====
num_taps = 1001
lpf = firwin(num_taps, bw, fs=fs, window="hamming")
y_rec = lfilter(lpf, 1, y_demod)
y_rec /= np.max(np.abs(y_rec))  # normalize


# ====================== PLAY + SAVE ======================
sd.play(y_rec, int(fs)); sd.wait()

out = file_path.replace("_QNT", "_DEMOD")
write(out, int(fs), (y_rec*32767).astype(np.int16))

print("\n>> Đã lưu file:", out)


# ====================== FFT & SPECTROGRAM ======================
Y1 = np.fft.fft(y_q)
Y2 = np.fft.fft(y_rec)
freqs = np.fft.fftfreq(N, 1/fs)

# vẽ full 0 → fs/2
f_plot = fs
idx = np.where((freqs >= 0) & (freqs <= f_plot*2))

plot_fft(
    freqs[idx], Y1[idx], Y2[idx],
    fs=fs, cutoff=bw,  
    label1="Tín hiệu lượng tử",
    label2="Sau Demod",
    title=f"FFT AM DEMOD (0 → {f_plot:.0f} Hz)"
)

plot_spectrogram(
    y_q, y_rec, fs,
    fmax=f_plot,
    title1="Trước Demod",
    title2="Sau Demod",
    suptitle=f"Spectrogram Full-band — fs={fs}Hz | fc={fc}Hz"
)
