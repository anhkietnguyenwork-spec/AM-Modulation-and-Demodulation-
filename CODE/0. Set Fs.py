import numpy as np
import sounddevice as sd
import os
from tkinter import Tk, filedialog
from wav_reader import read_wav
from dsp_utils import plot_fft, plot_spectrogram
import soundfile as sf 
from openpyxl import Workbook, load_workbook

# ==================== CHỌN FILE WAV ====================
root = Tk(); root.withdraw()
file_path = filedialog.askopenfilename(
    title="Chọn file WAV",
    filetypes=[("File WAV","*.wav")]
)
if not file_path: exit("LỖI: CHƯA CHỌN FILE!")

file_name = os.path.splitext(os.path.basename(file_path))[0]
xlsx_path = file_name + ".xlsx"
sheet_name = "Code2025"


# ==================== ĐỌC WAV & INPUT FFT ====================
initial_fs = int(input("Nhập Fs của file (Hz) = "))

# ==================== GHI EXCEL ====================
if not os.path.exists(xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"], ws["B1"] = "Fs_original", "FFT_Cutoff"
else:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

ws["A2"] = initial_fs

wb.save(xlsx_path)
print(f"📊 Đã lưu bảng Excel → {xlsx_path} (A2,B2)")

