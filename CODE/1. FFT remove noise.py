import numpy as np
import sounddevice as sd
import os
from tkinter import Tk, filedialog
from wav_reader import read_wav
from dsp_utils import plot_fft, plot_spectrogram
import soundfile as sf       
from openpyxl import Workbook, load_workbook


# ==================== CHỌN FILE WAV ====================
root = Tk(); root.withdraw()
file_path = filedialog.askopenfilename(
    title="Chọn file WAV",
    filetypes=[("File WAV","*.wav")]
)
if not file_path: exit("LỖI: CHƯA CHỌN FILE!")

file_name = os.path.splitext(os.path.basename(file_path))[0]
xlsx_path = file_name + ".xlsx"
sheet_name = "Code2025"



# ==================== TÊN FILE EXCEL TƯƠNG ỨNG ====================
xlsx_path = file_path.replace(".wav", ".xlsx")   # Cùng tên – đổi đuôi
if not os.path.exists(xlsx_path):
    exit(f"LỖI: Không tìm thấy file Excel: {xlsx_path}")


# ==================== ĐỌC THÔNG TIN TỪ EXCEL ====================
wb = load_workbook(xlsx_path)
ws = wb.active
initial_fs = ws["A2"].value                         # ⬅ lấy Fs từ FFT
if initial_fs is None:
    exit("LỖI: Chưa có Fs => Phải chạy file ""0.Set Fs.py"" trước!")

y, fs = read_wav(file_path, initial_fs)
if y.ndim>1: y = y[:,0]

y = y/np.max(np.abs(y))
N = len(y)

cutoff = float(input("Tần số lọc nhiễu FFT (Hz) = "))
f_plot_max = float(input("Vẽ phổ FFT đến tần số (Hz) = "))
f_plot_max *= 2 


# ==================== FFT xử lý ====================
Y  = np.fft.fft(y)
freqs = np.fft.fftfreq(N,1/fs)

Yf = Y.copy()
Yf[np.abs(freqs)>cutoff] = 0

yfft = np.fft.ifft(Yf).real
yfft = yfft/np.max(np.abs(yfft))


# ==================== Playback lựa chọn ====================
print("\n--------------- PHÁT ÂM THANH ----------------")
print("1 = Nghe tín hiệu gốc")
print("2 = Nghe tín hiệu đã lọc FFT")
print("3 = Nghe lần lượt cả hai")
print("0 = Bỏ qua phát thử")
choice = input("Chọn cách phát âm thanh: ")

if choice=="1": 
    print("Phát tín hiệu gốc...")
    sd.play(y,fs); sd.wait()

elif choice=="2": 
    print("Phát tín hiệu sau lọc FFT...")
    sd.play(yfft,fs); sd.wait()

elif choice=="3":
    print("Phát tín hiệu gốc...")
    sd.play(y,fs); sd.wait()
    print("Phát tín hiệu sau lọc...")
    sd.play(yfft,fs); sd.wait()

print("-------------------------------------------------\n")


# ==================== Lưu WAV ====================
out_audio = file_name + "_FFT.wav"
sf.write(out_audio, yfft.astype("float32"), fs)
print("\n>> Đã lưu output:", out_audio)


# ==================== GHI EXCEL ====================
if not os.path.exists(xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["B1"] = "FFT_Cutoff"
else:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

ws["B2"] = cutoff
wb.save(xlsx_path)

print(f">> Đã ghi Excel → {xlsx_path} (A2,B2)")


# ==================== VẼ ====================
idx = np.where((freqs>=0)&(freqs<=f_plot_max))

plot_fft(freqs[idx],Y[idx],Yf[idx],
         fs=fs,
         cutoff=cutoff,
         label1="Phổ gốc",
         label2=f"Phổ sau lọc LPF: {cutoff}Hz",
         title=f"FFT — 0 → {f_plot_max} Hz (x2 range)")


plot_spectrogram(
    y, yfft, fs,
    fmax=f_plot_max/2,
    title1="Phổ gốc",
    title2="Phổ sau lọc FFT",
    suptitle=f"SPECTROGRAM"
)
