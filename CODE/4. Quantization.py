import numpy as np
import sounddevice as sd
import os
from tkinter import Tk, filedialog
from wav_reader import read_wav
from scipy.io.wavfile import write
from dsp_utils import plot_fft, plot_spectrogram
import matplotlib.pyplot as plt
from openpyxl import load_workbook  # <<< cần để lấy ô E2 & ghi F2


# ==================== CHỌN FILE ====================
root = Tk(); root.withdraw()
file_path = filedialog.askopenfilename(
    title="Chọn file tín hiệu sau Sampling (_SAMP.wav)",
    filetypes=[("WAV", "*_SAMP.wav")]
)
if not file_path:
    exit("LỖI: Chưa chọn file!")

# ================== LẤY Fs_new TỪ FILE EXCEL ==================
xlsx = file_path.replace("_SAMP.wav", ".xlsx")
if not os.path.exists(xlsx):
    exit("LỖI: Không tìm thấy Excel đi kèm file!")

wb = load_workbook(xlsx)
ws = wb.active

initial_fs = ws["E2"].value 
if initial_fs is None:
    exit("LỖI: Ô A2 rỗng → chưa thực hiện Sampling nên chưa có Fs!")

print(f"\nFs từ Excel (A2) = {initial_fs} Hz → dùng để lượng tử hoá")


# ==================== ĐỌC WAV (GIỮ ĐÚNG Fs SAMPLED) ====================
y, fs_new = read_wav(file_path, int(initial_fs))
if y.ndim > 1: y = y[:,0]

y = y.astype(float)
y /= np.max(np.abs(y))
N = len(y)

print(f"Đã đọc xong file — Fs' = {fs_new} Hz | Số mẫu = {N}")


# ==================== INPUT BIT + FFT RANGE ====================
bits       = int(input("Số bit lượng tử = "))
f_plot_max = float(input("Vẽ phổ FFT đến (Hz) = "))

if bits < 1:
    exit("LỖI: bits phải ≥ 1")


# =============== GHI SỐ BIT VÀO Ô F2 =================
ws["F2"] = bits
wb.save(xlsx)
print(f">> Đã ghi F2 = {bits}-bit vào Excel\n")


# ==================== QUANTIZATION ====================
levels = 2**(bits+1)
delta  = 2/(levels-1)

y_q = np.round((y+1)/delta)*delta - 1
y_q = np.clip(y_q,-1,1)

print(f">> Lượng tử hóa {bits}-bit → {levels/2} mức.")


# ==================== PLAY + SAVE ====================
sd.play(y_q, fs_new); sd.wait()

out = file_path.replace("_SAMP","_QNT")
write(out, fs_new, (y_q*32767).astype(np.int16))

print("\n>> Đã lưu:", out)

# ==================== ĐỒ THỊ THEO MIỀN THỜI GIAN ====================
t=np.arange(N)/fs_new
Tview=0.001

plt.figure(figsize=(10,4))
plt.plot(t[t<=Tview],y[t<=Tview],label="Trước lượng tử",lw=1.4)
plt.step(t[t<=Tview],y_q[t<=Tview],where='mid',
         label=f"Sau lượng tử ({bits}-bit)",lw=1.3)

plt.title("Time Domain — Analog Sample vs Digital Quantized")
plt.xlabel("Time (s)"); plt.ylabel("Amplitude")
plt.grid(True); plt.legend(); plt.tight_layout()
plt.show()
